import pickle
import os.path
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from google.auth.transport.requests import Request
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
from openpyxl import load_workbook

from object_defs import QuestionEntryObject
from object_defs import DriveFileObject
from object_defs import CategoryObject

class DriveInterface():
    def __init__(self):
        self.service = None
        self.files = []

    def authenticate(self):
        # Authenticate
        SCOPES = ['https://www.googleapis.com/auth/drive']
        creds = None
        if os.path.exists('token.pickle'):
            with open('token.pickle', 'rb') as token:
                creds = pickle.load(token)
        # - if not valid, remove token.pickle and authenticate again
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
                creds = flow.run_local_server(port=0)
                with open('token.pickle', 'wb') as token:
                    pickle.dump(creds, token)

        self.service = build('drive', 'v3', credentials=creds)


    #List files with an xlsx extension and return them in a list
    def list_files_with_extension(self, extension):
        results = self.service.files().list(pageSize=10, fields="nextPageToken, files(id, name)").execute()
        items = results.get('files', [])

        if not items:
            print('No files found.')
        else:
            for item in items:
                if item['name'].endswith(extension):
                    self.files.append(DriveFileObject.DriveFileObject(item['name'], item['id']))

        return self.files
    
    #Download File
    def download_file_with_name(self, file_name):
        request = self.service.files().get_media(fileId=self.get_file_id_with_name(file_name))
        with open('downloaded_file.xlsx', 'wb') as fh:
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while done is False:
                status, done = downloader.next_chunk()
                print(f"Download {int(status.progress() * 100)}%.")

    #Parse File
    def parse_file(self):
        #for now, default to loading this 'downloaded_file.xlsx', better solution later
        wb = load_workbook('downloaded_file.xlsx')
        color_guide = wb.worksheets[0]
        questions = wb.worksheets[1]
        categoryDict = {}
        #print all entries in color_guide
        for row in color_guide.iter_rows():
            for cell in row:
                print(cell.value + " : " + str(cell.fill.start_color.index))
                categoryDict[str(cell.fill.start_color.index)] = cell.value

        #read in questions and create a new QuestionEntryObject for each
        for row in questions.iter_rows():
           questionEntry = QuestionEntryObject.QuestionEntryObject(row[0].value,    #question
                                               categoryDict[str(row[0].fill.start_color.index)],    #question_category
                                               row[1].value,    #question_type
                                               row[2].value,    #question_points
                                               row[3].value,    #question_time
                                               row[4].value,    #media_url
                                               row[5].value,    #note
                                               row[6].value,    #link
                                               row[7].value,    #correct_answer
                                               row[8].value,    #choice2
                                               row[9].value,    #choice3
                                               row[10].value,   #choice4
                                               row[11].value,   #etc
                                               row[12].value,   #unused
                                               row[13].value,   #zone1
                                               row[14].value,   #zone2
                                               row[15].value,   #zone3
                                               row[16].value,   #zone4
                                               row[17].value)





    def get_file_id_with_name(self, file_name):
        for file in self.files:
            if file.name == file_name:
                return file.id
        return None

